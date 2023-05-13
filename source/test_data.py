import os.path
import numpy as np
from nistrng import *
import csv
from openpyxl import Workbook, load_workbook
import multiprocessing as mp
from source import create_battery_of_tests


def test_prep(binary_sequence):
    all_test_dict = create_battery_of_tests.create_all_battery()
    eligible_battery = create_eligible_battery(binary_sequence, all_test_dict)

    return all_test_dict, eligible_battery


def create_eligible_battery(binary_sequence: np.ndarray, all_battery):
    # Check the eligibility of the test and generate an eligible battery from the default NIST-sp800-22r1a battery
    eligible_battery: dict = check_eligibility_all_battery(binary_sequence, all_battery)

    return eligible_battery


def test_data(binary_sequence: np.ndarray, eligible_battery: dict, check_eligibility: bool = True) -> []:
    num_processes = min(mp.cpu_count(), len(eligible_battery))  # determine the number of available processes
    pool = mp.Pool(processes=num_processes)  # create a process pool
    # results = pool.starmap(run_by_name_battery, [(name, sequence, eligible_battery, check_eligibility) for name in
    # eligible_battery.keys()])
    results = []
    for name in eligible_battery.keys():
        sequence = binary_sequence.copy()  # create a copy of the binary sequence for this test
        result = pool.apply_async(run_by_name_battery, (name, sequence, eligible_battery, check_eligibility))
        results.append(result)

    pool.close()  # close the process pool
    pool.join()  # wait for all processes to complete
    # return results
    return [r.get() for r in results]


def save_results_csv(results, filename):
    file_exists = os.path.isfile(filename)
    with open(filename, "a", newline="") as csvfile:
        fieldnames = ["name", "score", "result", "time"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        for result, elapsed_time in results:
            writer.writerow({"name": result.name, "score": np.round(result.score, 3),
                             "result": "passed" if result.passed else "failed", "time": str(elapsed_time) + "ms"})
    return filename


def save_results_xlsx(results, filename):
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Write headers for new sheet
        sheet["A1"] = "Test Name"
        sheet["B1"] = "Score"
        sheet["C1"] = "Result"
        sheet["D1"] = "Time (ms)"

    # Get the last used row in the sheet
    last_row = sheet.max_row

    # Add a blank row before new data
    sheet.insert_rows(last_row + 1)

    # Write results to sheet
    for i, (result, elapsed_time) in enumerate(results):
        sheet.cell(row=last_row + 2 + i, column=1, value=result.name)
        sheet.cell(row=last_row + 2 + i, column=2, value=round(result.score, 3))
        sheet.cell(row=last_row + 2 + i, column=3, value="Passed" if result.passed else "Failed")
        sheet.cell(row=last_row + 2 + i, column=4, value=elapsed_time)

    workbook.save(filename)
    return filename


def save_results_txt(results, filename):
    file_exists = os.path.isfile(filename)
    with open(filename, "a") as txtfile:
        if not file_exists:
            txtfile.write("{:<40} {:<10} {:<10} {:<10}\n".format("name", "score", "result", "time"))
        for result, elapsed_time in results:
            txtfile.write("{:<40} {:<10.3f} {:<10} {:<10}ms\n".format(result.name, result.score,
                                                                      "passed" if result.passed else "failed",
                                                                      str(elapsed_time)))
    return filename
